from openpyxl import load_workbook

wb = load_workbook('testimonials/google_forms_feedback.xlsx')
ws = wb.active

# Print all rows
for i, row in enumerate(ws.iter_rows(values_only=True)):
    print(f"\n=== Row {i} ===")
    for j, cell in enumerate(row):
        if cell is not None and str(cell).strip():
            print(f"Column {j}: {cell}")
